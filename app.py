import tkinter as tk
from tkinter import messagebox, simpledialog
import cv2
import face_recognition
import pickle
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ================= CONFIG =================

DB_FILE = "face_db.pkl"
EXCEL_FILE = "attendance.xlsx"
SAMPLES_REQUIRED = 20

# ================= DATABASE =================

def save_db(db):
    with open(DB_FILE, "wb") as f:
        pickle.dump(db, f)

def load_db():
    if not os.path.exists(DB_FILE):
        return {}

    with open(DB_FILE, "rb") as f:
        data = pickle.load(f)

    # New dict format
    if isinstance(data, dict):
        return data

    # Old tuple format → auto migrate
    if isinstance(data, tuple) and len(data) == 3:
        encodings, names, ids = data
        db = {}

        for enc, name, reg_id in zip(encodings, names, ids):
            if reg_id not in db:
                db[reg_id] = {
                    "name": name,
                    "encoding": enc,
                    "samples": 1,
                    "enrolled_on": "unknown"
                }
            else:
                db[reg_id]["encoding"] = (
                    db[reg_id]["encoding"] + enc
                ) / 2
                db[reg_id]["samples"] += 1

        save_db(db)
        return db

    return {}

# ================= ENROLLMENT =================

def enroll_student():
    name = simpledialog.askstring("Enroll", "Enter Name")
    if not name:
        return

    reg_id = simpledialog.askstring("Enroll", "Enter Registration ID")
    if not reg_id:
        return

    db = load_db()

    if reg_id in db:
        overwrite = messagebox.askyesno(
            "Already Exists",
            "This Registration ID already exists.\nDo you want to overwrite?"
        )
        if not overwrite:
            return

    cap = cv2.VideoCapture(0)
    collected = []

    while len(collected) < SAMPLES_REQUIRED:
        ret, frame = cap.read()
        if not ret:
            continue

        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        locations = face_recognition.face_locations(rgb)

        if len(locations) == 1:
            enc = face_recognition.face_encodings(rgb, locations)[0]
            collected.append(enc)

            top, right, bottom, left = locations[0]
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

        cv2.putText(
            frame,
            f"Samples: {len(collected)}/{SAMPLES_REQUIRED}",
            (20, 40),
            cv2.FONT_HERSHEY_SIMPLEX,
            1,
            (0, 255, 0),
            2
        )

        cv2.imshow("Enrollment (Press Q to cancel)", frame)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            cap.release()
            cv2.destroyAllWindows()
            return

    cap.release()
    cv2.destroyAllWindows()

    avg_encoding = sum(collected) / len(collected)

    db[reg_id] = {
        "name": name,
        "encoding": avg_encoding,
        "samples": SAMPLES_REQUIRED,
        "enrolled_on": datetime.now().strftime("%Y-%m-%d")
    }

    save_db(db)
    messagebox.showinfo("Success", f"{name} enrolled successfully")

# ================= POPUP =====================
def show_recognition_popup(name, reg_id):
    popup = tk.Toplevel(root)
    popup.overrideredirect(True)  # no title bar
    popup.configure(bg="#111827")

    # Position popup center-top
    popup.geometry("300x80+{}+{}".format(
        root.winfo_x() + 110,
        root.winfo_y() + 60
    ))

    tk.Label(
        popup,
        text="Attendance Marked",
        font=("Segoe UI", 10, "bold"),
        bg="#111827",
        fg="#22c55e"
    ).pack(pady=(8, 0))

    tk.Label(
        popup,
        text=f"{name}\n{reg_id}",
        font=("Segoe UI", 11),
        bg="#111827",
        fg="white"
    ).pack()

    # Auto close after 2 seconds
    popup.after(2000, popup.destroy)


# ================= ATTENDANCE =================

def start_attendance():
    db = load_db()
    if not db:
        messagebox.showerror("Error", "No students enrolled")
        return

    today_file = datetime.now().strftime("%Y-%m-%d") + ".xlsx"

    # Create Excel if it doesn't exist
    if not os.path.exists(today_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name", "Registration ID", "Time"])
        wb.save(today_file)

    wb = load_workbook(today_file)
    ws = wb.active

    already_marked = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[1]:
            already_marked.add(row[1])

    known_encodings = []
    known_ids = []
    known_names = []

    for reg_id, info in db.items():
        known_encodings.append(info["encoding"])
        known_ids.append(reg_id)
        known_names.append(info["name"])

    marked_now = set()
    buffer = []

    cap = cv2.VideoCapture(0)

    while True:
        ret, frame = cap.read()
        if not ret:
            continue

        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        locations = face_recognition.face_locations(rgb)
        encodings = face_recognition.face_encodings(rgb, locations)

        for enc, loc in zip(encodings, locations):
            matches = face_recognition.compare_faces(
                known_encodings, enc, tolerance=0.6
            )

            top, right, bottom, left = loc

            if True in matches:
                idx = matches.index(True)
                reg_id = known_ids[idx]
                name = known_names[idx]

                if reg_id not in already_marked and reg_id not in marked_now:
                    time_now = datetime.now().strftime("%H:%M:%S")
                    buffer.append((name, reg_id, time_now))
                    marked_now.add(reg_id)

                    # ✅ show confirmation popup
                    show_recognition_popup(name, reg_id)


                cv2.rectangle(frame, (left, top), (right, bottom), (0, 180, 0), 2)
                cv2.putText(
                    frame,
                    name,
                    (left, top - 10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.7,
                    (0, 180, 0),
                    2
                )
            else:
                cv2.rectangle(frame, (left, top), (right, bottom), (200, 0, 0), 2)
                cv2.putText(
                    frame,
                    "Unknown",
                    (left, top - 10),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.7,
                    (200, 0, 0),
                    2
                )

        cv2.imshow("Attendance (Press Q to stop)", frame)
        if cv2.waitKey(1) & 0xFF == ord("q"):
            break

    cap.release()
    cv2.destroyAllWindows()

    for name, reg_id, time in buffer:
        ws.append([name, reg_id, time])

    wb.save(today_file)
    wb.close()



# ================= DELETE =================

def delete_student():
    reg_id = simpledialog.askstring("Delete", "Enter Registration ID")
    if not reg_id:
        return

    db = load_db()
    if reg_id not in db:
        messagebox.showerror("Error", "Registration ID not found")
        return

    confirm = messagebox.askyesno(
        "Confirm Delete",
        f"Delete {db[reg_id]['name']} ({reg_id})?"
    )
    if not confirm:
        return

    del db[reg_id]
    save_db(db)
    messagebox.showinfo("Deleted", "Student removed successfully")

# ================= VIEW STUDENTS =============

def view_registered_students():
    db = load_db()

    if not db:
        messagebox.showinfo("Registered Students", "No students registered.")
        return

    viewer = tk.Toplevel(root)
    viewer.title("Registered Students")
    viewer.geometry("500x400")
    viewer.resizable(False, False)

    tk.Label(
        viewer,
        text="Registered Students",
        font=("Arial", 14, "bold")
    ).pack(pady=10)

    frame = tk.Frame(viewer)
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side="right", fill="y")

    listbox = tk.Listbox(
        frame,
        yscrollcommand=scrollbar.set,
        font=("Consolas", 11),
        width=60,
        height=15
    )
    listbox.pack(side="left", fill="both", expand=True)

    scrollbar.config(command=listbox.yview)

    # Header
    listbox.insert("end", f"{'Reg ID':<15} | Name")
    listbox.insert("end", "-" * 45)

    # Sorted by Reg ID
    for reg_id in sorted(db.keys()):
        name = db[reg_id]["name"]
        listbox.insert("end", f"{reg_id:<15} | {name}")


# ================= OPEN EXCEL =================

def open_excel():
    if not os.path.exists(EXCEL_FILE):
        messagebox.showerror("Error", "attendance.xlsx not found")
        return
    os.startfile(EXCEL_FILE)


# ================= UI =================

BG_COLOR = "#f4f6f9"
CARD_COLOR = "#ffffff"
PRIMARY_COLOR = "#1f7ae0"
TEXT_COLOR = "#111827"
SUBTEXT_COLOR = "#6b7280"

root = tk.Tk()
root.title("Face Recognition Attendance System")
root.geometry("520x620")
root.configure(bg=BG_COLOR)
root.resizable(False, False)

# ---------- Card Container ----------
card = tk.Frame(
    root,
    bg=CARD_COLOR,
    bd=0,
    relief="flat"
)
card.place(relx=0.5, rely=0.5, anchor="center", width=440, height=560)

# ---------- Title ----------
tk.Label(
    card,
    text="Face Recognition Attendance",
    font=("Segoe UI", 18, "bold"),
    bg=CARD_COLOR,
    fg=TEXT_COLOR
).pack(pady=(30, 5))

tk.Label(
    card,
    text="Foundation in Data Science • Attendance System",
    font=("Segoe UI", 10),
    bg=CARD_COLOR,
    fg=SUBTEXT_COLOR
).pack(pady=(0, 25))

# ---------- Button Style ----------
def styled_button(parent, text, command):
    return tk.Button(
        parent,
        text=text,
        command=command,
        font=("Segoe UI", 11, "bold"),
        bg=PRIMARY_COLOR,
        fg="white",
        activebackground="#1e40af",
        activeforeground="white",
        bd=0,
        relief="flat",
        cursor="hand2",
        height=2
    )

# ---------- Buttons ----------
styled_button(card, "Enroll Student", enroll_student).pack(fill="x", padx=40, pady=8)
styled_button(card, "Start Attendance", start_attendance).pack(fill="x", padx=40, pady=8)
styled_button(card, "View Registered Students", view_registered_students).pack(fill="x", padx=40, pady=8)
styled_button(card, "Delete Student", delete_student).pack(fill="x", padx=40, pady=8)
styled_button(card, "Open Attendance Excel", open_excel).pack(fill="x", padx=40, pady=8)

# ---------- Divider ----------
tk.Frame(card, bg="#e5e7eb", height=1).pack(fill="x", padx=40, pady=20)

# ---------- Exit ----------
tk.Button(
    card,
    text="Exit Application",
    command=root.destroy,
    font=("Segoe UI", 10),
    bg="#e5e7eb",
    fg=TEXT_COLOR,
    bd=0,
    relief="flat",
    cursor="hand2",
    height=2
).pack(fill="x", padx=40)

# ---------- Footer ----------
tk.Label(
    card,
    text="Designed for academic use • Offline • Secure",
    font=("Segoe UI", 9),
    bg=CARD_COLOR,
    fg=SUBTEXT_COLOR
).pack(side="bottom", pady=15)

root.mainloop()
