import cv2
import face_recognition
import pickle
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

DB_FILE = "face_db.pkl"
EXCEL_FILE = "attendance.xlsx"

# Load face database
with open(DB_FILE, "rb") as f:
    known_encodings, known_names, known_ids = pickle.load(f)

# Create or load Excel file
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Registration ID", "Date", "Time"])
    wb.save(EXCEL_FILE)

wb = load_workbook(EXCEL_FILE)
ws = wb.active

marked_ids = set()

cap = cv2.VideoCapture(0)

while True:
    ret, frame = cap.read()
    if not ret:
        continue

    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    locations = face_recognition.face_locations(rgb)
    encodings = face_recognition.face_encodings(rgb, locations)

    for encoding, loc in zip(encodings, locations):
        matches = face_recognition.compare_faces(
            known_encodings, encoding, tolerance=0.5
        )

        name = "Unknown"
        reg_id = ""

        if True in matches:
            index = matches.index(True)
            name = known_names[index]
            reg_id = known_ids[index]

            if reg_id not in marked_ids:
                now = datetime.now()
                ws.append([
                    name,
                    reg_id,
                    now.strftime("%Y-%m-%d"),
                    now.strftime("%H:%M:%S")
                ])
                wb.save(EXCEL_FILE)
                marked_ids.add(reg_id)

        top, right, bottom, left = loc
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)
        cv2.putText(
            frame,
            name,
            (left, top - 10),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.8,
            (0, 255, 0),
            2
        )

    cv2.imshow("Attendance System - Press Q to Exit", frame)

    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

cap.release()
cv2.destroyAllWindows()
wb.close()
